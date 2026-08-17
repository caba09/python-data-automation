from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURATION
# ============================================================

PROJECT_ROOT = Path(__file__).resolve().parent.parent

REPORTS_DIR = PROJECT_ROOT / "reports"

OUTPUT_FILE = (
    REPORTS_DIR / "trippulse_business_report.xlsx"
)


# ============================================================
# DATA LOADING
# ============================================================

def load_report_data() -> dict[str, pd.DataFrame]:
    """Load analytics and quality reports."""

    return {
        "kpis": pd.read_csv(
            REPORTS_DIR / "business_kpis.csv"
        ),
        "revenue_by_segment": pd.read_csv(
            REPORTS_DIR / "revenue_by_segment.csv"
        ),
        "revenue_by_destination": pd.read_csv(
            REPORTS_DIR / "revenue_by_destination.csv"
        ),
        "bookings_by_destination": pd.read_csv(
            REPORTS_DIR / "bookings_by_destination.csv"
        ),
        "revenue_by_channel": pd.read_csv(
            REPORTS_DIR / "revenue_by_channel.csv"
        ),
        "monthly_revenue": pd.read_csv(
            REPORTS_DIR / "monthly_revenue.csv"
        ),
        "data_quality": pd.read_csv(
            REPORTS_DIR / "data_quality_report.csv"
        ),
    }


# ============================================================
# EXCEL WRITING
# ============================================================

def write_excel_report(
    data: dict[str, pd.DataFrame],
) -> None:
    """Create the Excel workbook."""

    REPORTS_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl",
    ) as writer:

        data["kpis"].to_excel(
            writer,
            sheet_name="Executive Summary",
            index=False,
        )

        data["revenue_by_destination"].to_excel(
            writer,
            sheet_name="Destinations",
            index=False,
        )

        data["revenue_by_segment"].to_excel(
            writer,
            sheet_name="Customers",
            index=False,
        )

        data["monthly_revenue"].to_excel(
            writer,
            sheet_name="Monthly Revenue",
            index=False,
        )

        data["revenue_by_channel"].to_excel(
            writer,
            sheet_name="Channels",
            index=False,
        )

        data["data_quality"].to_excel(
            writer,
            sheet_name="Data Quality",
            index=False,
        )


# ============================================================
# GENERAL FORMATTING
# ============================================================

def format_headers(worksheet) -> None:
    """Format worksheet headers."""

    for cell in worksheet[1]:

        cell.font = Font(
            bold=True,
        )

        cell.alignment = Alignment(
            horizontal="center",
        )


def autofit_columns(worksheet) -> None:
    """Automatically adjust column widths."""

    for column_cells in worksheet.columns:

        max_length = 0

        column_letter = get_column_letter(
            column_cells[0].column
        )

        for cell in column_cells:

            if cell.value is not None:

                max_length = max(
                    max_length,
                    len(str(cell.value)),
                )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max_length + 3,
            45,
        )


# ============================================================
# WORKBOOK FORMATTING
# ============================================================

def format_workbook() -> None:
    """Apply general workbook formatting."""

    workbook = load_workbook(
        OUTPUT_FILE
    )

    for worksheet in workbook.worksheets:

        worksheet.freeze_panes = "A2"

        format_headers(
            worksheet
        )

        autofit_columns(
            worksheet
        )

    # --------------------------------------------------------
    # Executive Summary
    # --------------------------------------------------------

    worksheet = workbook[
        "Executive Summary"
    ]

    worksheet.column_dimensions["A"].width = 30
    worksheet.column_dimensions["B"].width = 25

    for row in range(
        2,
        worksheet.max_row + 1,
    ):

        metric = worksheet.cell(
            row=row,
            column=1,
        ).value

        value_cell = worksheet.cell(
            row=row,
            column=2,
        )

        if metric in {
            "total_revenue",
            "average_booking_value",
        }:

            value_cell.number_format = (
                '€#,##0.00'
            )

        elif metric in {
            "cancellation_rate",
        }:

            value_cell.number_format = (
                '0.00"%"'
            )

        elif metric in {
            "total_bookings",
            "unique_customers",
        }:

            value_cell.number_format = (
                '#,##0'
            )

        elif metric in {
            "average_trip_duration",
        }:

            value_cell.number_format = (
                '0.00" days"'
            )

    # --------------------------------------------------------
    # Data Quality
    # --------------------------------------------------------

    worksheet = workbook[
        "Data Quality"
    ]

    worksheet.column_dimensions["A"].width = 25
    worksheet.column_dimensions["B"].width = 40
    worksheet.column_dimensions["C"].width = 15
    worksheet.column_dimensions["D"].width = 15

    for row in range(
        2,
        worksheet.max_row + 1,
    ):

        status = worksheet.cell(
            row=row,
            column=4,
        ).value

        if status == "FAIL":

            for cell in worksheet[row]:

                cell.font = Font(
                    bold=True,
                )

    workbook.save(
        OUTPUT_FILE
    )


# ============================================================
# CHART DATA
# ============================================================

def create_chart_data_sheet(
    workbook,
    monthly_revenue: pd.DataFrame,
) -> None:
    """Create a hidden sheet containing chart-ready data."""

    worksheet = workbook.create_sheet(
        "Chart Data"
    )

    # --------------------------------------------------------
    # Monthly revenue
    # --------------------------------------------------------

    worksheet["A1"] = "Metric"

    worksheet["A2"] = "Revenue"

    for cell in worksheet[1]:

        cell.font = Font(
            bold=True,
        )

    for index, row in monthly_revenue.iterrows():

        column = index + 2

        worksheet.cell(
            row=1,
            column=column,
            value=row["month"],
        )

        worksheet.cell(
            row=2,
            column=column,
            value=row["total_amount"],
        )

    # Hide technical sheet.
    worksheet.sheet_state = "hidden"


# ============================================================
# CHARTS
# ============================================================
def add_charts() -> None:
    """Add business charts to the workbook."""

    workbook = load_workbook(
        OUTPUT_FILE
    )

    # ========================================================
    # MONTHLY REVENUE
    # ========================================================

    worksheet = workbook[
        "Monthly Revenue"
    ]

    # Remove existing charts.
    worksheet._charts = []

    chart = LineChart()

    chart.title = "Monthly Revenue"
    chart.y_axis.title = "Revenue (€)"
    chart.x_axis.title = "Month"

    chart.height = 9
    chart.width = 18

    # Revenue values: B2:B19
    values = Reference(
        worksheet,
        min_col=2,
        min_row=2,
        max_row=worksheet.max_row,
    )

    # Months: A2:A19
    categories = Reference(
        worksheet,
        min_col=1,
        min_row=2,
        max_row=worksheet.max_row,
    )

    # Create exactly one series.
    series = Series(
        values,
        title="Revenue",
    )

    chart.series = [
        series
    ]

    chart.set_categories(
        categories
    )

    # One metric = no legend needed.
    chart.legend = None

    worksheet.add_chart(
        chart,
        "D2",
    )

    # ========================================================
    # TOP DESTINATIONS
    # ========================================================

    worksheet = workbook[
        "Destinations"
    ]

    worksheet._charts = []

    chart = BarChart()

    chart.type = "bar"

    chart.title = (
        "Top 10 Destinations by Revenue"
    )

    chart.x_axis.title = "Revenue (€)"
    chart.y_axis.title = "Destination"

    chart.height = 10
    chart.width = 18

    max_row = min(
        worksheet.max_row,
        11,
    )

    values = Reference(
        worksheet,
        min_col=4,
        min_row=1,
        max_row=max_row,
    )

    categories = Reference(
        worksheet,
        min_col=2,
        min_row=2,
        max_row=max_row,
    )

    chart.add_data(
        values,
        titles_from_data=True,
    )

    chart.set_categories(
        categories
    )

    # One metric = no legend needed.
    chart.legend = None

    worksheet.add_chart(
        chart,
        "F2",
    )

    workbook.save(
        OUTPUT_FILE
    )
# ============================================================
# MAIN
# ============================================================

def main() -> None:
    """Generate the final TripPulse business report."""

    print(
        "Loading report data..."
    )

    data = load_report_data()

    print(
        "Creating Excel workbook..."
    )

    write_excel_report(
        data
    )

    print(
        "Formatting workbook..."
    )

    format_workbook()

    print(
        "Adding charts..."
    )

    add_charts()

    print(
        "\nReport created successfully:"
        f"\n{OUTPUT_FILE}"
    )


if __name__ == "__main__":
    main()